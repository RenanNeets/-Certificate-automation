#Importa a biblioteca de usar planilha
import openpyxl
#importando 2° biblioteca
from PIL import Image, ImageDraw, ImageFont

#Abrindo a planilha e guardando em var

workbookAlunos = openpyxl.load_workbook('planilha_alunos.xlsx')
sheetAlunos = workbookAlunos['Sheet1']

#Indo em uma linha e pegando todas as células dela
for indice, linha in enumerate(sheetAlunos.iter_rows(min_row=2, max_row=2)): #Para limitar a geração usa o "iter_rows(min_row=2,max_row=3)"
    nomeCurso = linha[0].value
    nomeParticipante = linha[1].value
    tipoParticipacao = linha[2].value
    dataInico = linha[3].value
    dataFinal = linha[4].value
    cargaHoraria = linha[5].value
    dataEmissao = linha[6].value

    #Transferindo os dados da planilha pra imagem
    fonteNegrito = ImageFont.truetype('./TAHOMABD.TTF', 90)
    fonteGeral = ImageFont.truetype('./TAHOMA.TTF', 80)
    fonteData = ImageFont.truetype('./TAHOMA.TTF', 55)

    image = Image.open('./certificado_padrao.jpg')
    desenhar = ImageDraw.Draw(image)
                  # X , Y  ;  variavel       ; cor dela    ;  fonte
    desenhar.text((1020,827), nomeParticipante, fill='black', font=fonteNegrito)
    desenhar.text((1060,950), nomeCurso, fill='black', font=fonteGeral)
    desenhar.text((1435,1065), tipoParticipacao, fill='black', font=fonteGeral)
    desenhar.text((1480,1182), str(cargaHoraria), fill='black', font=fonteGeral)
    #Datas
    desenhar.text((750,1770), dataInico, fill='black', font=fonteData)
    desenhar.text((750,1930), dataFinal, fill='black', font=fonteData)
    desenhar.text((2220,1930), dataEmissao, fill='black', font=fonteData)
    



    image.save(f'./{indice} {nomeParticipante} certificado.png') #Salvando o arquivo
